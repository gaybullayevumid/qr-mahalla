from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

from .models import Region, District, Mahalla
from apps.houses.models import House


class ExportHousesView(APIView):
    """
    Export houses data to Excel.

    Query parameters:
    - region: Region ID (optional)
    - district: District ID (optional)
    - mahalla: Mahalla ID (optional)

    Permission:
    - Admin: Can export all data
    - Leader: Can export only their mahalla data
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        region_id = request.query_params.get("region")
        district_id = request.query_params.get("district")
        mahalla_id = request.query_params.get("mahalla")

        # Check permissions
        if user.role == "leader":
            # Leader can only export their own mahalla
            if not hasattr(user, "mahalla"):
                return Response(
                    {"error": "Leader must be assigned to a mahalla"},
                    status=status.HTTP_403_FORBIDDEN,
                )

            # Force mahalla_id to user's mahalla
            mahalla_id = user.mahalla.id

            # If leader tries to specify different filters, deny
            if region_id or district_id:
                return Response(
                    {"error": "Leader can only export their own mahalla data"},
                    status=status.HTTP_403_FORBIDDEN,
                )

        elif user.role != "admin":
            return Response(
                {"error": "Only admin and leader can export data"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Build queryset based on filters
        houses = House.objects.select_related(
            "owner", "mahalla__district__region"
        ).all()

        # Apply filters
        if mahalla_id:
            houses = houses.filter(mahalla_id=mahalla_id)
            try:
                mahalla = Mahalla.objects.select_related("district__region").get(
                    id=mahalla_id
                )
                export_name = f"{mahalla.district.region.name}_{mahalla.district.name}_{mahalla.name}"
            except Mahalla.DoesNotExist:
                return Response(
                    {"error": "Mahalla not found"}, status=status.HTTP_404_NOT_FOUND
                )
        elif district_id:
            houses = houses.filter(mahalla__district_id=district_id)
            try:
                district = District.objects.select_related("region").get(id=district_id)
                export_name = f"{district.region.name}_{district.name}"
            except District.DoesNotExist:
                return Response(
                    {"error": "District not found"}, status=status.HTTP_404_NOT_FOUND
                )
        elif region_id:
            houses = houses.filter(mahalla__district__region_id=region_id)
            try:
                region = Region.objects.get(id=region_id)
                export_name = region.name
            except Region.DoesNotExist:
                return Response(
                    {"error": "Region not found"}, status=status.HTTP_404_NOT_FOUND
                )
        else:
            export_name = "All_Data"

        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Xonadonlar"

        # Header style
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            "ID",
            "Viloyat",
            "Tuman",
            "Mahalla",
            "Manzil",
            "Uy raqami",
            "Egasining telefoni",
            "Egasining ismi",
            "Egasining familiyasi",
            "Yaratilgan sana",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Data rows
        for row_idx, house in enumerate(houses, 2):
            ws.cell(row=row_idx, column=1, value=house.id)
            ws.cell(row=row_idx, column=2, value=house.mahalla.district.region.name)
            ws.cell(row=row_idx, column=3, value=house.mahalla.district.name)
            ws.cell(row=row_idx, column=4, value=house.mahalla.name)
            ws.cell(row=row_idx, column=5, value=house.address)
            ws.cell(row=row_idx, column=6, value=house.house_number)
            ws.cell(
                row=row_idx, column=7, value=house.owner.phone if house.owner else ""
            )
            ws.cell(
                row=row_idx,
                column=8,
                value=house.owner.first_name if house.owner else "",
            )
            ws.cell(
                row=row_idx,
                column=9,
                value=house.owner.last_name if house.owner else "",
            )
            ws.cell(
                row=row_idx,
                column=10,
                value=house.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            )

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Prepare response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Xonadonlar_{export_name}_{timestamp}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response
